from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic import TemplateView
from django.views.generic.list import ListView
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.db.models import Q
from django.db.models import Prefetch
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse
import datetime
import calendar
from collections import defaultdict

from booking.models import Booking, BookingChannel, Payment, Enquiry, Expense
from property.models import Property, PropertyChannel, PropertyBlockDate
from tasks.models import Task
from django.db.models import Sum, Count
from django.utils import timezone
from django.db.models.functions import ExtractYear, ExtractMonth
from django.utils.safestring import mark_safe
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.mail import send_mail, EmailMessage, get_connection
from django.conf import settings
from django.views.decorators.http import require_POST
from .utils import calculateTotalPayment
from accounts.utils import get_visible_user_ids
from accounts.models import CoHost
from django.contrib.staticfiles.storage import staticfiles_storage

class HostDashboardAPIView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/host/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_type = self.request.GET.get('type', '') 
        # print('selected_type==>',selected_type)
        context['selected_type'] = selected_type
        context['active_listings_count'] = Property.objects.filter(created_by__in=get_visible_user_ids(self.request.user), status='Active').count()

        total_revenue_agg = Payment.objects.filter(
            booking__property__created_by__in=get_visible_user_ids(self.request.user),
            is_paid=True
        ).exclude(
            type__in=['Refundable deposit', 'Refundable to guest']
        ).aggregate(total=Sum('amount'))
        context['total_revenue'] = total_revenue_agg['total'] or 0

        context['pending_tasks_count'] = Task.objects.filter(created_by__in=get_visible_user_ids(self.request.user), completed=False).count()
        context['new_enquiry_list'] = Enquiry.objects.filter(property__created_by__in=get_visible_user_ids(self.request.user), is_archive=False, is_booked=False)[:6]

        today = self.request.user.get_local_date()

        if selected_type == 'tomorrow':
            target_date = today + datetime.timedelta(days=1)
        else:
            target_date = today

        active_bookings = Booking.objects.filter(
            property__created_by__in=get_visible_user_ids(self.request.user), 
            property__status='Active'
        ).exclude(status='cancelled')

        total_checkin = active_bookings.filter(check_in_date=target_date).count()
        total_checkout = active_bookings.filter(check_out_date=target_date).count()
        total_tasks = Task.objects.filter(
            created_by__in=get_visible_user_ids(self.request.user),
            date=target_date,
            completed=False
        ).count()
        
        total_payment = Payment.objects.filter(
            booking__property__created_by__in=get_visible_user_ids(self.request.user),
            expected_payment_date__date=target_date,
            is_paid=False
        ).exclude(
            type__in = ['Refundable deposit', 'Cleaning Fee', 'Application Fee', 'Taxes', 'Other Fees', 'Refundable to guest']
        ).count()

        # Determine which properties are currently occupied by guests
        inhouse_property_ids = set(active_bookings.filter(
            check_in_date__lte=target_date,
            check_out_date__gte=target_date
        ).values_list('property_id', flat=True))

        # Determine which properties are manually blocked on this date
        blocked_property_ids = set(PropertyBlockDate.objects.filter(
            property__created_by__in=get_visible_user_ids(self.request.user),
            property__status='Active',
            is_active=True,
            start_date__lte=target_date,
            end_date__gte=target_date
        ).values_list('property_id', flat=True))

        total_inhouse = len(inhouse_property_ids)
        total_occupied = len(inhouse_property_ids | blocked_property_ids)
        total_active_count = Property.objects.filter(created_by__in=get_visible_user_ids(self.request.user), status='Active').count()
        total_vacant = total_active_count - total_occupied


        context['total_checkin'] = total_checkin
        context['total_checkout'] = total_checkout
        context['total_tasks'] = total_tasks
        context['total_payment'] = total_payment
        context['total_inhouse'] = total_inhouse
        context['total_vacant'] = total_vacant

        seven_days_ago = today - datetime.timedelta(days=7)
        context['recent_bookings'] = Booking.objects.filter(
            property__created_by__in=get_visible_user_ids(self.request.user),
            created_at__date__gte=seven_days_ago,
            created_at__date__lte=today
        ).exclude(status='cancelled').order_by('-created_at')

        current_year = timezone.now().year

        # 1) Determine selected year (query param ?year=YYYY, fallback to current year)
        selected_year_str = self.request.GET.get("year")
        try:
            selected_year = int(selected_year_str) if selected_year_str else current_year
        except ValueError:
            selected_year = current_year

        gross_qs = Booking.objects.filter(
            property__created_by__in=get_visible_user_ids(self.request.user),
            check_in_date__year=selected_year
        )
        context['gross_bookings_amount'] = gross_qs.aggregate(total=Sum('price'))['total'] or 0
        context['gross_bookings_count'] = gross_qs.aggregate(count=Count('id'))['count'] or 0
        channel_data = gross_qs.values('channel__name').annotate(
            count=Count('id'),
            total=Sum('price')
        ).order_by('-count')
        context['channel_labels'] = [c['channel__name'] or 'Direct' for c in channel_data]
        context['channel_counts'] = [c['count'] for c in channel_data]
        context['channel_amounts'] = [float(c['total'] or 0) for c in channel_data]
        context['channel_labels_json'] = mark_safe(json.dumps([c['channel__name'] or 'Direct' for c in channel_data]))
        context['channel_counts_json'] = mark_safe(json.dumps([c['count'] for c in channel_data]))
        context['channel_amounts_json'] = mark_safe(json.dumps([float(c['total'] or 0) for c in channel_data]))

        # --- Earning monthly for a year ---
        
        # print( 'current_year==>',current_year)
        # print( 'selected_year==>',selected_year)


        # 2) Month‑wise earnings from Payment for ALL years (only paid ones)
        qs = (
            Payment.objects
            .filter(
                booking__property__created_by__in=get_visible_user_ids(self.request.user),
                is_paid=True,
            )
            .exclude(type__in=['Refundable deposit', 'Refundable to guest'])
            .annotate(
                year=ExtractYear('expected_payment_date'),
                month=ExtractMonth('expected_payment_date'),
            )
            .values('year', 'month')
            .annotate(total=Sum('amount'))
            .order_by('year', 'month')
        )

        # print('qs==>',qs)

        # 3) Build {year: [12 month totals]} using SUM from query
        year_earnings = {}
        for row in qs:
            # print('row==>',row)
            y = row['year']
            m = row['month']      # 1–12
            if y is None or m is None:
                continue

            if y not in year_earnings:
                year_earnings[y] = [0] * 12

            # this is the dynamic sum for that month in that year
            year_earnings[y][m - 1] = float(row['total'])

        # 4) If selected_year has no payments, fall back to current_year
        if selected_year not in year_earnings and current_year in year_earnings:
            selected_year = current_year

        # 5) Final 12‑length list for the selected year
        monthly_earnings = year_earnings.get(selected_year, [0] * 12)


        # total_earnings = sum(monthly_earnings)
        # monthly_average = total_earnings / 12

        # Monthly average of earnings elapsed so far:
        # - past year: plain average over all 12 months
        # - future year: nothing has elapsed yet
        # - current year: (revenue up to and including the current month /
        #   days elapsed since Jan 1) * 30, so numerator and denominator
        #   always cover the same period
        year_start = datetime.date(selected_year, 1, 1)
        today_date = today

        if selected_year < today_date.year:
            monthly_average = sum(monthly_earnings) / 12
        elif selected_year > today_date.year:
            monthly_average = 0
        else:
            total_earnings = sum(monthly_earnings[:today_date.month])
            days_elapsed = (today_date - year_start).days + 1  # include today
            monthly_average = (total_earnings / days_elapsed) * 30
        

        # 6) Years for dropdown — use values_list + set for SQLite compatibility
        total_years = sorted(set(
            Booking.objects
            .filter(property__created_by__in=get_visible_user_ids(self.request.user))
            .exclude(status='cancelled')
            .values_list('check_in_date__year', flat=True)
        ))

        context['current_year'] = current_year
        context['selected_year'] = selected_year
        context['total_years'] = total_years
        context['year_earnings_json'] = mark_safe(json.dumps(year_earnings))
        context['monthly_earnings_js'] = mark_safe(json.dumps(monthly_earnings))
        context['monthly_average'] = format(monthly_average, '.2f')

        # --- Upcoming events (check-in, check-out, payment, tasks) grouped by date ---
        upcoming_month_str = self.request.GET.get("upcoming_month")
        upcoming_year_str = self.request.GET.get("upcoming_year")

        try:
            upcoming_month = int(upcoming_month_str) if upcoming_month_str else today.month
            if upcoming_month < 1 or upcoming_month > 12:
                upcoming_month = today.month
        except ValueError:
            upcoming_month = today.month

        try:
            upcoming_year = int(upcoming_year_str) if upcoming_year_str else today.year
        except ValueError:
            upcoming_year = today.year

        payments_qs = Payment.objects.filter(
            Q(expected_payment_date__year=upcoming_year, expected_payment_date__month=upcoming_month) |
            Q(payment_date__year=upcoming_year, payment_date__month=upcoming_month)
        )

        bookings = (
            Booking.objects
            .filter(
                Q(check_in_date__year=upcoming_year, check_in_date__month=upcoming_month) |
                Q(check_out_date__year=upcoming_year, check_out_date__month=upcoming_month),
                property__created_by__in=get_visible_user_ids(self.request.user)
            )
            .exclude(status='cancelled')
            .select_related('property', 'channel')
            .prefetch_related('images', Prefetch('payments', queryset=payments_qs))
        )

        tasks = (
            Task.objects
            .filter(
                created_by__in=get_visible_user_ids(self.request.user),
                property__created_by__in=get_visible_user_ids(self.request.user),
                date__gte=today,
                date__year=upcoming_year,
                date__month=upcoming_month,
                completed=False,
            )
            .select_related('property')
        )

        def status_for(d):
            diff = (d - today).days
            if diff < 0:
                return ("overdue", "Overdue")
            if diff == 0:
                return ("due-soon", "Due today")
            if diff <= 7:
                return ("due-soon", f"Due in {diff} days")
            return ("upcoming", f"Due in {diff} days")

        def get_property_image(prop, fallback=""):
            getter = getattr(prop, "get_primary_image_url", None)
            if callable(getter):
                try:
                    url = getter()
                except Exception:
                    url = ""
                if url:
                    return url
            return fallback

        event_groups = defaultdict(list)
        group_sort = {"Checkin": 1, "Checkout": 2, "Payment": 3, "Task": 4}

        for b in bookings:
            prop_img = b.images.filter(is_primary=True).first() or b.images.first()
            avatar = prop_img.image.url if prop_img else staticfiles_storage.url('img/common/default_user_icon.png')
            guest_name = f"{b.first_name or ''} {b.last_name or ''}".strip() or "Guest"
            property_image = get_property_image(b.property, avatar)
            guest_phone = " ".join([p for p in [b.country_code, b.phone] if p])
            guest_email = b.email or ""
            channel_name = b.channel.name if b.channel else ""
            stay_str = ""
            if b.check_in_date and b.check_out_date:
                nights = (b.check_out_date - b.check_in_date).days
                stay_str = f"{b.check_in_date:%b %d, %Y} - {b.check_out_date:%b %d, %Y} ({nights} nights)"

            if b.check_in_date and b.check_in_date.year == upcoming_year and b.check_in_date.month == upcoming_month:
                code, label = status_for(b.check_in_date)
                event_groups[b.check_in_date].append({
                    "date": b.check_in_date,
                    "group": "Checkin",
                    "title": "Check In",
                    "property": b.property.title,
                    "icon": "/img/property/checkin.png",
                    "guest_name": guest_name,
                    "guest_phone": guest_phone,
                    "guest_email": guest_email,
                    "stay": stay_str,
                    "channel_name": channel_name,
                    "avatar": avatar,
                    "property_image": property_image,
                    "status_code": code,
                    "status_label": label,
                    "status_class": "status-checkin",
                    "sort_key": group_sort["Checkin"],
                })

            if b.check_out_date and b.check_out_date.year == upcoming_year and b.check_out_date.month == upcoming_month:
                code, label = status_for(b.check_out_date)
                event_groups[b.check_out_date].append({
                    "date": b.check_out_date,
                    "group": "Checkout",
                    "title": "Check Out",
                    "property": b.property.title,
                    "icon": "img/property/checkout.png",
                    "guest_name": guest_name,
                    "guest_phone": guest_phone,
                    "guest_email": guest_email,
                    "stay": stay_str,
                    "channel_name": channel_name,
                    "avatar": avatar,
                    "property_image": property_image,
                    "status_code": code,
                    "status_label": label,
                    "status_class": "status-checkout",
                    "sort_key": group_sort["Checkout"],
                })

            for p in b.payments.all():
                if p.type in ['Refundable deposit', 'Cleaning Fee', 'Application Fee', 'Taxes', 'Other Fees', 'Refundable to guest']:
                    continue
                pay_date = None
                if p.expected_payment_date:
                    pay_date = p.expected_payment_date.date()
                elif p.payment_date:
                    pay_date = p.payment_date.date()
                if not pay_date:
                    continue
                if pay_date.year != upcoming_year or pay_date.month != upcoming_month:
                    continue
                code, label = status_for(pay_date)
                if p.is_paid:
                    code, label = "paid", "Paid"
                event_groups[pay_date].append({
                    "date": pay_date,
                    "group": "Payment",
                    "title": "Payment",
                    "amount": float(p.amount),
                    "paid": p.is_paid,
                    "property": b.property.title,
                    "icon": "img/task/type/payment.svg",
                    "guest_name": guest_name,
                    "guest_phone": guest_phone,
                    "guest_email": guest_email,
                    "stay": stay_str,
                    "channel_name": channel_name,
                    "avatar": avatar,
                    "property_image": property_image,
                    "status_code": code,
                    "status_label": label,
                    "status_class": "status-payment",
                    "sort_key": group_sort["Payment"],
                })

        for t in tasks:
            if not t.date:
                continue
            code, label = status_for(t.date)
            title = "Task"
            status_class = "status-checkin"
            if t.task_type == "cleaning":
                title = "Cleaning"
                status_class = "status-cleaning"
            elif t.task_type == "maintenance":
                title = "Maintenance"
                status_class = "status-checkout"
            elif t.task_type == "payment":
                title = "Payment Task"
                status_class = "status-payment"

            t_prop_image = get_property_image(t.property, "")
            event_groups[t.date].append({
                "date": t.date,
                "group": "Task",
                "title": title,
                "property": t.property.title,
                "task_type": t.task_type,
                "icon": {
                    "cleaning": "img/task/type/cleaning.svg",
                    "maintenance": "img/task/type/maintenance.svg",
                    "payment": "img/task/type/payment.svg",
                }.get(t.task_type, "img/task/type/others.svg"),
                "guest_name": t.assigned_to or "Task",
                "guest_phone": t.phone or "",
                "guest_email": "",
                "stay": "",
                "channel_name": "",
                "avatar": staticfiles_storage.url('img/common/default_user_icon_1.png'),
                "property_image": t_prop_image,
                "status_code": "completed" if t.completed else code,
                "status_label": "Completed" if t.completed else label,
                "status_class": status_class,
                "sort_key": group_sort["Task"],
            })

        upcoming_groups = []
        for group_date in sorted(event_groups.keys()):
            if group_date >= today:
                group_code, group_label = status_for(group_date)
                items = sorted(event_groups[group_date], key=lambda e: e["sort_key"])
                upcoming_groups.append({
                    "date": group_date,
                    "status_label": group_label,
                    "status_code": group_code,
                    "items": items,
                })

        context["upcoming_month"] = upcoming_month
        context["upcoming_year"] = upcoming_year
        context["upcoming_groups"] = upcoming_groups
        return context


class RevenueByListingView(LoginRequiredMixin, TemplateView):
    """Per-listing revenue analytics for a selected year.

    Mirrors the design at revenue-dashboard.html but is wired to live data.
    All monetary figures are stored/aggregated in USD (the platform base) and
    rendered in the viewer's display currency via the {% money %} tag.

    Revenue is attributed on an *accrual / per-night* basis: each non-cancelled
    booking's net price (price minus the refundable deposit for manual bookings,
    matching the calendar logic) is spread evenly across its nights, and each
    night is counted in the month it falls in. This keeps Days booked,
    Occupancy %, Average Daily rate and Revenue internally consistent.
    """
    template_name = 'frontend/revenue/revenue_by_listing.html'

    MONTH_LABELS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Donut / channel colour palette (cycled when there are many channels).
    PALETTE = ['#ef4444', '#3b82f6', '#facc15', '#22c55e', '#f59e0b',
               '#b91c1c', '#ec4899', '#8b5cf6', '#06b6d4', '#14b8a6',
               '#a855f7', '#cbd5e1']

    def dispatch(self, request, *args, **kwargs):
        # Co-hosts do not have access to the accountant (revenue) section.
        if request.user.is_authenticated and CoHost.objects.filter(co_host=request.user).exists():
            messages.error(request, 'Revenue is not available for co-hosts.')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        """Serve a file export when ?export=xlsx|csv, otherwise the HTML page."""
        export = request.GET.get('export')
        if export == 'xlsx':
            return self._export_xlsx(request)
        if export == 'csv':
            return self._export_csv(request)
        return super().get(request, *args, **kwargs)

    def _export_csv(self, request):
        """Build a clean, Excel-ready CSV from the same computed report.

        Money is converted to the viewer's display currency and written as
        plain numbers (no symbol/grouping) so spreadsheets treat them as
        numeric. A UTF-8 BOM is prepended so symbols render correctly in Excel.
        """
        import csv
        import io
        import re
        from accounts import currency as cur

        context = self.get_context_data()
        code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()
        symbol = cur.symbol_for(code)
        months = list(context.get('month_labels', self.MONTH_LABELS))
        sel = context.get('selected_property')
        year = context.get('selected_year')
        listing_title = sel['title'] if sel else 'All listings'

        try:
            generated = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
        except Exception:
            generated = timezone.now().strftime('%Y-%m-%d %H:%M')

        buf = io.StringIO()
        buf.write('﻿')  # UTF-8 BOM so Excel renders ₹/€/₨ etc. correctly
        writer = csv.writer(buf)

        # --- Metadata header ---
        writer.writerow(['FavHost - Revenue by Listing'])
        writer.writerow(['Listing', listing_title])
        writer.writerow(['Year', year])
        writer.writerow(['Currency', f'{code} ({symbol})'])
        writer.writerow(['Generated', generated])
        writer.writerow([])

        def cell(value, fmt):
            # Money -> numeric in display currency; pct/int pass through as-is.
            return cur.money_raw(value, code) if fmt == 'money' else value

        def section(title, first_header, rows, empty_message=None):
            writer.writerow([title])
            writer.writerow([first_header] + months + ['Overall'])
            if rows:
                for r in rows:
                    label = r['label'].replace(' $', '')
                    writer.writerow(
                        [label]
                        + [cell(v, r['fmt']) for v in r['values']]
                        + [cell(r['overall'], r['fmt'])]
                    )
            elif empty_message:
                writer.writerow([empty_message])
            writer.writerow([])

        if sel and context.get('has_data'):
            section('Monthly Performance', 'Metric', context['table_rows'])
            section('Number of bookings by Channel', 'Channel',
                    context['bookings_channel_rows'], f'No bookings for {year}.')
            section(f'Revenue by Channel ({code})', 'Channel',
                    context['revenue_channel_rows'], f'No channel revenue for {year}.')
        else:
            writer.writerow(['No data available for this selection.'])

        safe = re.sub(r'[^A-Za-z0-9._-]+', '_', listing_title).strip('_') or 'listing'
        filename = f'revenue_{safe}_{year}.csv'

        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _export_xlsx(self, request):
        """Build a styled, Excel-native .xlsx so column widths/number formats
        are preserved — labels like 'Average Daily rate' show in full, and
        money is written as real numbers in the viewer's display currency.
        """
        import io
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from accounts import currency as cur

        context = self.get_context_data()
        code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()
        symbol = cur.symbol_for(code)
        months = list(context.get('month_labels', self.MONTH_LABELS))
        sel = context.get('selected_property')
        year = context.get('selected_year')
        listing_title = sel['title'] if sel else 'All listings'
        try:
            generated = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
        except Exception:
            generated = timezone.now().strftime('%Y-%m-%d %H:%M')

        ncols = 1 + 12 + 1  # Metric/Channel + 12 months + Overall

        # --- Styles (mirror the updated on-screen palette: brand orange
        #     accents, charcoal section bands, warm-orange overall cells) ---
        title_font = Font(bold=True, size=14, color='EB5310')
        meta_label_font = Font(bold=True, color='6B7280')
        section_font = Font(bold=True, color='FFFFFF')
        section_fill = PatternFill('solid', fgColor='313131')
        header_font = Font(bold=True, color='1A1A1A')
        header_fill = PatternFill('solid', fgColor='FAF7F5')
        overall_fill = PatternFill('solid', fgColor='FFF3EC')
        overall_font = Font(bold=True, color='C2410C')
        metric_font = Font(bold=True, color='1A1A1A')
        metric_fill = PatternFill('solid', fgColor='FAF7F5')
        empty_font = Font(italic=True, color='6B7280')
        thin = Side(style='thin', color='E5E7EB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        MONEY_FMT = '#,##0.00'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Revenue'

        r = 1
        ws.cell(r, 1, 'FavHost - Revenue by Listing').font = title_font
        r += 1
        for label, val in [('Listing', listing_title), ('Year', year),
                           ('Currency', f'{code} ({symbol})'), ('Generated', generated)]:
            ws.cell(r, 1, label).font = meta_label_font
            ws.cell(r, 2, val)
            r += 1
        r += 1  # spacer

        def write_section(title, first_header, rows, empty_message=None):
            nonlocal r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            tc = ws.cell(r, 1, title)
            tc.font = section_font
            tc.fill = section_fill
            tc.alignment = Alignment(horizontal='center')
            r += 1
            for i, h in enumerate([first_header] + months + ['Overall'], start=1):
                hc = ws.cell(r, i, h)
                hc.font = overall_font if h == 'Overall' else header_font
                hc.fill = overall_fill if h == 'Overall' else header_fill
                hc.border = border
                hc.alignment = Alignment(horizontal='left' if i == 1 else 'center')
            r += 1
            if rows:
                for row in rows:
                    lc = ws.cell(r, 1, row['label'].replace(' $', ''))
                    lc.font = metric_font
                    lc.fill = metric_fill
                    lc.border = border
                    for i, v in enumerate(list(row['values']) + [row['overall']], start=2):
                        cell = ws.cell(r, i)
                        if row['fmt'] == 'money':
                            cell.value = float(cur.money_raw(v, code))
                            cell.number_format = MONEY_FMT
                        elif row['fmt'] == 'pct':
                            cell.value = float(v)
                            cell.number_format = '0.00'
                        else:
                            cell.value = v
                        cell.border = border
                        cell.alignment = Alignment(horizontal='right')
                        if i == ncols:
                            cell.fill = overall_fill
                            cell.font = overall_font
                    r += 1
            elif empty_message:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
                ws.cell(r, 1, empty_message).font = empty_font
                r += 1
            r += 1  # spacer after section

        if sel and context.get('has_data'):
            write_section('Monthly Performance', 'Metric', context['table_rows'])
            write_section('Number of bookings by Channel', 'Channel',
                          context['bookings_channel_rows'], f'No bookings for {year}.')
            write_section(f'Revenue by Channel ({code})', 'Channel',
                          context['revenue_channel_rows'], f'No channel revenue for {year}.')
        else:
            ws.cell(r, 1, 'No data available for this selection.').font = empty_font

        # Column widths: wide first column so labels show in full; months even.
        ws.column_dimensions['A'].width = 22
        for i in range(2, ncols):
            ws.column_dimensions[get_column_letter(i)].width = 11
        ws.column_dimensions[get_column_letter(ncols)].width = 12  # Overall

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe = re.sub(r'[^A-Za-z0-9._-]+', '_', listing_title).strip('_') or 'listing'
        filename = f'revenue_{safe}_{year}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_ids = get_visible_user_ids(self.request.user)
        current_year = timezone.now().year

        # ---- Listings for the sidebar (all of the viewer's properties) ----
        properties = list(
            Property.objects.filter(created_by__in=user_ids).order_by('title')
        )
        listings = [{
            'id': str(p.id),
            'title': p.title,
            'image': p.get_primary_image_url(),
        } for p in properties]

        # ---- Resolve the selected listing ----
        listing_id = self.request.GET.get('listing')
        selected = None
        if listing_id:
            selected = next((p for p in properties if str(p.id) == str(listing_id)), None)
        if selected is None and properties:
            selected = properties[0]

        # ---- Years available for the dropdown (from this listing's bookings) ----
        # Include both check-in and check-out years so a stay that spans a year
        # boundary (e.g. checks in Dec, out in Jan) makes the later year
        # selectable — otherwise the revenue that spills into it is unreachable.
        if selected is not None:
            year_rows = (
                Booking.objects
                .filter(property=selected)
                .exclude(status='cancelled')
                .values_list('check_in_date__year', 'check_out_date__year')
            )
            years = sorted({
                y for pair in year_rows for y in pair if y
            })
        else:
            years = []
        if current_year not in years:
            years = sorted(set(years) | {current_year})

        try:
            selected_year = int(self.request.GET.get('year') or current_year)
        except (TypeError, ValueError):
            selected_year = current_year
        if selected_year not in years:
            selected_year = current_year

        _yidx = years.index(selected_year) if selected_year in years else -1
        context.update({
            'listings': listings,
            'selected_property': {'id': str(selected.id), 'title': selected.title} if selected else None,
            'total_years': years,
            'selected_year': selected_year,
            'year_prev': years[_yidx - 1] if _yidx > 0 else None,
            'year_next': years[_yidx + 1] if 0 <= _yidx < len(years) - 1 else None,
            'current_year': current_year,
            'month_labels': self.MONTH_LABELS,
        })

        # ---- No listings yet: render an empty shell ----
        if selected is None:
            context.update({
                'has_data': False,
                'table_rows': [],
                'bookings_channel_rows': [],
                'revenue_channel_rows': [],
                'bar_cols': [],
                'donut_segments': [],
                'donut_gradient': 'conic-gradient(#e5e7eb 0% 100%)',
            })
            return context

        # ---- Monthly accumulators (index 0 = Jan .. 11 = Dec) ----
        days_in_month = [calendar.monthrange(selected_year, m)[1] for m in range(1, 13)]
        inquiries = [0] * 12
        checkins = [0] * 12
        checkouts = [0] * 12
        cancellations = [0] * 12
        tasks_arr = [0] * 12
        days_booked = [0] * 12
        revenue = [0.0] * 12
        bookings_by_channel = defaultdict(lambda: [0] * 12)
        revenue_by_channel = defaultdict(lambda: [0.0] * 12)

        refundable = float(selected.refundable_deposit or 0)

        bookings = Booking.objects.filter(property=selected).select_related('channel')
        for b in bookings:
            ci, co = b.check_in_date, b.check_out_date
            channel_name = b.channel.name if b.channel else 'Direct'
            cancelled = b.status == 'cancelled'

            # Count-based rows keyed on check-in / check-out month.
            if ci and ci.year == selected_year:
                if cancelled:
                    cancellations[ci.month - 1] += 1
                else:
                    checkins[ci.month - 1] += 1
                    bookings_by_channel[channel_name][ci.month - 1] += 1
            if co and co.year == selected_year and not cancelled:
                checkouts[co.month - 1] += 1

            # Revenue / nights attribution (non-cancelled stays only).
            if cancelled or not ci or not co or co <= ci:
                continue
            net_price = float(b.price or 0)
            if not b.external_uid and refundable:
                net_price = max(net_price - refundable, 0.0)
            total_nights = (co - ci).days
            per_night = net_price / total_nights if total_nights else 0.0

            night = ci
            while night < co:
                if night.year == selected_year:
                    m = night.month - 1
                    days_booked[m] += 1
                    revenue[m] += per_night
                    revenue_by_channel[channel_name][m] += per_night
                night += datetime.timedelta(days=1)

        # Inquiries by requested check-in month.
        for ci in Enquiry.objects.filter(
            property=selected, check_in_date__year=selected_year
        ).values_list('check_in_date', flat=True):
            if ci:
                inquiries[ci.month - 1] += 1

        # Tasks scheduled in the year.
        for d in Task.objects.filter(
            property=selected, date__year=selected_year
        ).values_list('date', flat=True):
            if d:
                tasks_arr[d.month - 1] += 1

        # ---- Derived rows: occupancy + average daily rate ----
        occupancy = [
            round(days_booked[m] / days_in_month[m] * 100, 2) if days_in_month[m] else 0
            for m in range(12)
        ]
        adr = [
            (revenue[m] / days_booked[m]) if days_booked[m] else 0.0
            for m in range(12)
        ]

        total_days = sum(days_in_month)
        total_booked = sum(days_booked)
        total_revenue = sum(revenue)
        overall_occ = round(total_booked / total_days * 100, 2) if total_days else 0
        overall_adr = (total_revenue / total_booked) if total_booked else 0.0

        def fmt_pct(values):
            return ["%.2f" % v for v in values]

        context['table_rows'] = [
            {'label': 'Days in the month', 'values': days_in_month, 'overall': total_days, 'fmt': 'int'},
            {'label': 'Of Inquiries', 'values': inquiries, 'overall': sum(inquiries), 'fmt': 'int'},
            {'label': 'Checkins', 'values': checkins, 'overall': sum(checkins), 'fmt': 'int'},
            {'label': 'Checkouts', 'values': checkouts, 'overall': sum(checkouts), 'fmt': 'int'},
            {'label': 'Cancellations', 'values': cancellations, 'overall': sum(cancellations), 'fmt': 'int'},
            {'label': 'Tasks', 'values': tasks_arr, 'overall': sum(tasks_arr), 'fmt': 'int'},
            {'label': 'Days booked', 'values': days_booked, 'overall': total_booked, 'fmt': 'int'},
            {'label': 'Occupancy %', 'values': fmt_pct(occupancy), 'overall': "%.2f" % overall_occ, 'fmt': 'pct'},
            {'label': 'Average Daily rate', 'values': adr, 'overall': overall_adr, 'fmt': 'money'},
            {'label': 'Revenue $', 'values': revenue, 'overall': total_revenue, 'fmt': 'money'},
        ]

        # ---- Channel breakdown tables ----
        channel_names = sorted(set(bookings_by_channel) | set(revenue_by_channel))
        context['bookings_channel_rows'] = [{
            'label': name,
            'values': bookings_by_channel.get(name, [0] * 12),
            'overall': sum(bookings_by_channel.get(name, [0] * 12)),
            'fmt': 'int',
        } for name in channel_names]
        context['revenue_channel_rows'] = [{
            'label': name,
            'values': revenue_by_channel.get(name, [0.0] * 12),
            'overall': sum(revenue_by_channel.get(name, [0.0] * 12)),
            'fmt': 'money',
        } for name in channel_names]

        # ---- Bar chart: revenue by month (heights relative to peak month) ----
        max_rev = max(revenue) if revenue else 0
        context['bar_cols'] = [{
            'label': self.MONTH_LABELS[m],
            'value': revenue[m],
            'height': round(revenue[m] / max_rev * 100, 1) if max_rev > 0 else 0,
            'highlight': max_rev > 0 and revenue[m] == max_rev,
        } for m in range(12)]

        # ---- Donut: revenue share by channel for the year ----
        channel_year = sorted(
            ((name, sum(revenue_by_channel.get(name, [0.0] * 12))) for name in channel_names),
            key=lambda x: -x[1],
        )
        channel_year = [(n, a) for n, a in channel_year if a > 0]
        total_channel_rev = sum(a for _, a in channel_year)
        segments = []
        cum = 0.0
        for i, (name, amount) in enumerate(channel_year):
            pct = (amount / total_channel_rev * 100) if total_channel_rev else 0
            color = self.PALETTE[i % len(self.PALETTE)]
            segments.append({
                'name': name,
                'amount': amount,
                'pct': round(pct, 1),
                'color': color,
                'start': round(cum, 3),
                'end': round(cum + pct, 3),
            })
            cum += pct
        if segments:
            stops = ", ".join(f"{s['color']} {s['start']}% {s['end']}%" for s in segments)
            donut_gradient = f"conic-gradient({stops})"
        else:
            donut_gradient = "conic-gradient(#e5e7eb 0% 100%)"

        context['donut_segments'] = segments
        context['donut_gradient'] = donut_gradient
        context['has_data'] = True
        return context


@method_decorator(never_cache, name='dispatch')
class CalenderAPIView(LoginRequiredMixin,ListView):
    model = Booking
    template_name = 'frontend/calender/calender.html'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return self.get_bookings_data(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        properties = Property.objects.filter(created_by__in=get_visible_user_ids(self.request.user), status='Active' )

        rooms_list = []
        for prop in properties:
            rooms_list.append({
                'id': str(prop.id),
                'title': prop.title,
                'image': prop.get_primary_image_url(),
                'location': prop.location_display(),
                'price_per_night': float(prop.price_per_night or 0),
                'guest': prop.guest,
                'minimum_nights': prop.minimum_booking,
                'basePrice': float(prop.price_per_night),
                'weekendMultiplier': 1.0, 
                'slug': prop.slug,
                'detail_url': f"/property/detail/{prop.slug}/",
            })
        context['rooms_data'] = rooms_list
        context['bookings_data'] = []  # Initially empty, will be loaded by JS

        return context

    def get_bookings_data(self, request):
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        try:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

        bookings_qs = Booking.objects.filter(
            property__created_by__in=get_visible_user_ids(request.user),
            check_in_date__lte=end_date,
            check_out_date__gte=start_date
        ).exclude(status='cancelled').select_related('property', 'channel').prefetch_related('images', 'payments')

        bookings_list = []
        for b in bookings_qs:
            guest_image = b.images.filter(is_primary=True).first() or b.images.first()
            guest_avatar_url = guest_image.image.url if guest_image else staticfiles_storage.url('img/common/default_user_icon.png')
            total_payment = calculateTotalPayment(b.id)

            if b.check_in_date and b.check_out_date:
                total_price = b.price
                # Deduct refundable deposit for manual bookings where it is included in the price
                if not b.external_uid and b.property.refundable_deposit:
                    total_price = total_price - b.property.refundable_deposit
                
                payment_dates = []
                payments_info = {}
                first_installment_amount = 0.0

                for p in b.payments.all():
                    if p.type in ['Refundable deposit', 'Cleaning Fee', 'Application Fee', 'Taxes', 'Other Fees', 'Payment 1']:
                        first_installment_amount += float(p.amount)
                    if p.type in ['Refundable deposit', 'Cleaning Fee', 'Application Fee', 'Taxes', 'Other Fees', 'Refundable to guest']:
                        continue
                    if p.expected_payment_date:
                        d_str = p.expected_payment_date.date().isoformat()
                        payment_dates.append(d_str)
                        payments_info[d_str] = {
                            'amount': float(p.amount),
                            'is_paid': p.is_paid,
                            'type': p.type or 'Payment'
                        }

                bookings_list.append({
                    'id': str(b.id),
                    'propertyId': str(b.property.id),
                    'price_per_night': float(b.price_per_night),
                    'totalPrice': float(total_price),
                    'totalPayment': float(total_payment),
                    'firstInstallment': float(first_installment_amount),
                    'start_date': b.check_in_date.isoformat(),
                    'end_date': b.check_out_date.isoformat(),
                    'guestName': f"{b.first_name or ''} {b.last_name or ''}".strip(),
                    'guestAvatar': guest_avatar_url,
                    'color': b.channel.color if b.channel else '#808080',
                    'channelIcon': b.channel.white_icon.url if b.channel and b.channel.white_icon else '',
                    'channelName': b.channel.name if b.channel else 'N/A',
                    'bookingId': b.booking_id or 'N/A',
                    'paymentDates': payment_dates,
                    'paymentsInfo': payments_info,
                    'check_in_time': (b.check_in_time or b.property.check_in_time).strftime('%H:%M') if (b.check_in_time or b.property.check_in_time) else None,
                    'check_out_time': (b.check_out_time or b.property.check_out_time).strftime('%H:%M') if (b.check_out_time or b.property.check_out_time) else None,
                })

        # Fetch blocked dates
        blocked_dates_qs = PropertyBlockDate.objects.filter(
            property__created_by__in=get_visible_user_ids(request.user),
            start_date__lte=end_date,
            end_date__gte=start_date,
            is_active=True
        ).select_related('property')

        for block in blocked_dates_qs:
            bookings_list.append({
                'id': f"block-{block.id}",
                'propertyId': str(block.property.id),
                'start_date': block.start_date.isoformat(),
                'end_date': block.end_date.isoformat(),
                'color': '#9ca3af',  # Gray color for blocked dates
                'type': 'blocked',
                'reason': block.reason
            })

        # Fetch tasks for the date range organized by date and property
        tasks_qs = Task.objects.filter(
            created_by__in=get_visible_user_ids(request.user),
            date__lte=end_date,
            date__gte=start_date,
        ).select_related('property').order_by('date', 'time')

        tasks_list = []
        for task in tasks_qs:
            tasks_list.append({
                'id': f"task-{task.id}",
                'propertyId': str(task.property.id),
                'date': task.date.isoformat() if task.date else None,
                'time': task.time.isoformat() if task.time else None,
                'task_type': task.task_type,
                'details': task.details or '',
                'assigned_to': task.assigned_to or '',
                'phone': task.phone or '',
                'completed': task.completed,
                'repeat': task.repeat,
                'repeat_till': task.repeat_till.isoformat() if task.repeat_till else None,
                'type': 'task',
                'color': {
                    'cleaning': '#10b981',
                    'maintenance': '#f59e0b',
                    'payment': '#3b82f6',
                    'other': '#8b5cf6'
                }.get(task.task_type, '#6b7280')
            })

        # Fetch enquiries for the date range
        enquiries_qs = Enquiry.objects.filter(
            property__created_by__in=get_visible_user_ids(request.user),
            check_in_date__lte=end_date,
            check_out_date__gte=start_date,
            is_archive=False
        ).select_related('property')

        enquiries_list = []
        for eq in enquiries_qs:
            enquiries_list.append({
                'id': f"enquiry-{eq.id}",
                'propertyId': str(eq.property.id),
                'start_date': eq.check_in_date.isoformat(),
                'end_date': eq.check_out_date.isoformat(),
                'startDateStr': eq.check_in_date.isoformat(),
                'endDateStr': eq.check_out_date.isoformat(),
                'guestName': f"{eq.first_name or ''} {eq.last_name or ''}".strip() or "Enquiry Guest",
                'guestAvatar': staticfiles_storage.url('img/common/default_user_icon_1.png'),
                'color': '#f59e0b',
                'type': 'enquiry',
                'phone': eq.phone or '',
                'email': eq.email or '',
                'adults': eq.adults,
                'children': eq.children,
                'pets': eq.pets,
                'notes': eq.notes_for_host or '',
                'unique_id': str(eq.unique_id),
            })

        return JsonResponse({
            'bookings': bookings_list,
            'tasks': tasks_list,
            'enquiries': enquiries_list
        })


class CalendarTimelineView(CalenderAPIView):
    template_name = 'frontend/calender/calender_timeline.html'

# views.py
from datetime import date
from django.views.generic import ListView
from django.http import JsonResponse
from django.db.models import Prefetch

class CalendarListView(LoginRequiredMixin, ListView):
    model = Booking
    template_name = 'frontend/calender/calender-list-view.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['current_month'] = date.today()
        return ctx

    def get_queryset(self):
        # base queryset in case you still use object_list in template
        return (
            Booking.objects
            .filter(property__created_by__in=get_visible_user_ids(self.request.user))
            .exclude(status='cancelled')
            .select_related('property', 'channel')
            .prefetch_related('images', 'payments')
        )

    def get(self, request, *args, **kwargs):
        # AJAX JSON for list view
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            month = int(request.GET.get('month', date.today().month))
            year = int(request.GET.get('year', date.today().year))
            today = date.today()

            # Fetch bookings where check-in OR check-out is in the given month/year
            bookings = (
                Booking.objects
                .filter(
                    Q(property__created_by__in=get_visible_user_ids(request.user)),
                    Q(check_in_date__year=year, check_in_date__month=month) |
                    Q(check_out_date__year=year, check_out_date__month=month)
                )
                .exclude(status='cancelled')
                .select_related('property', 'channel')
                .prefetch_related(
                    'images',
                    Prefetch('payments', queryset=Payment.objects.filter(
                        expected_payment_date__year=year, expected_payment_date__month=month
                    ))
                )
            )

            tasks = (
                Task.objects
                .filter(
                    created_by__in=get_visible_user_ids(request.user),
                    property__created_by__in=get_visible_user_ids(request.user),
                    date__year=year,
                    date__month=month,
                )
                .select_related('property')
            )

            def status_for(d):
                diff = (d - today).days
                if diff < 0:
                    return ("overdue", "Overdue")
                if diff == 0:
                    return ("due-soon", "Due today")
                if diff <= 7:
                    return ("due-soon", f"Due in {diff} days")
                return ("upcoming", f"Due in {diff} days")

            events = []

            def format_guest_phone(country_code, phone):
                phone = (phone or "").strip()
                code = (country_code or "").strip()
                if not phone:
                    return ""
                if not code:
                    return phone
                # Ignore textual country codes like "in" and keep the stored phone only.
                if code.isalpha():
                    return phone
                if not code.startswith("+"):
                    code = f"+{code}"
                return f"{code} {phone}"

            # booking-based events
            for b in bookings:
                prop_img = b.images.filter(is_primary=True).first() or b.images.first()
                avatar = prop_img.image.url if prop_img else ""
                guest_name = f"{b.first_name or ''} {b.last_name or ''}".strip() or "Guest"
                stay_str = ""
                if b.check_in_date and b.check_out_date:
                    nights = (b.check_out_date - b.check_in_date).days
                    stay_str = f"{b.check_in_date:%b %d, %Y}—{b.check_out_date:%b %d, %Y} ({nights} nights)"

                # property image (prefer property’s own image helper if you have one)
                property_image = getattr(b.property, "get_primary_image_url", None)
                if callable(property_image):
                    property_image = property_image()
                else:
                    # fallback to first booking image
                    property_image = avatar or ""

                # Check‑out
                if b.check_out_date and b.check_out_date.year == year and b.check_out_date.month == month:
                    code, label = status_for(b.check_out_date)
                    events.append({
                        "id": str(b.id),
                        "date": b.check_out_date.isoformat(),
                        "group": "Checkout",
                        "title": "Check Out",
                        "property": b.property.title,
                        "location": b.property.city,
                        "guest_name": guest_name,
                        "guest_email": b.email or "",
                        "guest_phone": format_guest_phone(b.country_code, b.phone),
                        "stay": stay_str,
                        "avatar": avatar,
                        "property_image": property_image, 
                        "status_code": code,
                        "status_label": label,
                    })

                # Check‑in
                if b.check_in_date and b.check_in_date.year == year and b.check_in_date.month == month:
                    code, label = status_for(b.check_in_date)
                    events.append({
                        "id": str(b.id),
                        "date": b.check_in_date.isoformat(),
                        "group": "Checkin",
                        "title": "Check In",
                        "property": b.property.title,
                        "location": b.property.city,
                        "guest_name": guest_name,
                        "guest_email": b.email or "",
                        "guest_phone": format_guest_phone(b.country_code, b.phone),
                        "stay": stay_str,
                        "avatar": avatar,
                        "property_image": property_image,
                        "status_code": code,
                        "status_label": label,
                    })

                # Payments (from Payment model)
                for p in b.payments.all():
                    if p.type in ['Refundable deposit', 'Cleaning Fee', 'Application Fee', 'Taxes', 'Other Fees', 'Refundable to guest']:
                        continue
                    pay_date = p.expected_payment_date.date() if p.expected_payment_date else (p.payment_date.date() if p.payment_date else today)
                    code, label = status_for(pay_date)
                    if p.is_paid:
                        code, label = "paid", "Paid"
                    events.append({
                        "id": str(b.id),
                        "date": pay_date.isoformat(),
                        "group": "Payment",
                        "title": "Payment",
                        "amount": float(p.amount),
                        "paid": p.is_paid,
                        "property": b.property.title,
                        "location": b.property.city,
                        "guest_name": guest_name,
                        "guest_email": b.email or "",
                        "guest_phone": format_guest_phone(b.country_code, b.phone),
                        "stay": stay_str,
                        "avatar": avatar,
                        "property_image": property_image,
                        "status_code": code,
                        "status_label": label,
                    })

            # Task events (from Task model)
            for t in tasks:
                if not t.date:
                    continue
                code, label = status_for(t.date)
                title = "Task"
                if t.task_type == "cleaning":
                    title = "Cleaning"
                elif t.task_type == "maintenance":
                    title = "Maintenance"
                elif t.task_type == "payment":
                    title = "Payment Task 💳"

                # property image for tasks
                t_prop_image = getattr(t.property, "get_primary_image_url", None)
                if callable(t_prop_image):
                    t_prop_image = t_prop_image()
                else:
                    t_prop_image = ""

                events.append({
                    "id": t.id,
                    "date": t.date.isoformat(),
                    "time": t.time.isoformat() if t.time else None,
                    "group": "Task",
                    "title": title,
                    "details": t.details or "",
                    "property": t.property.title,
                    "location": t.property.city,
                    "guest_name": t.assigned_to or "",
                    "guest_email": "",
                    "guest_phone": format_guest_phone(t.country_code, t.phone),
                    "guest_id": t.phone or "",
                    "avatar": "",
                    "property_image": t_prop_image,
                    "status_code": "completed" if t.completed else code,
                    "status_label": "Completed" if t.completed else label,
                })

            # Enquiry events
            enquiries = Enquiry.objects.filter(
                property__created_by__in=get_visible_user_ids(request.user),
                check_in_date__year=year,
                check_in_date__month=month,
                is_booked=False,
            ).select_related('property')
            for e in enquiries:
                guest_name = f"{e.first_name or ''} {e.last_name or ''}".strip() or "Guest"
                stay_str = ""
                if e.check_in_date and e.check_out_date:
                    nights = (e.check_out_date - e.check_in_date).days
                    stay_str = f"{e.check_in_date:%b %d, %Y}—{e.check_out_date:%b %d, %Y} ({nights} nights)"
                events.append({
                    "id": str(e.unique_id),
                    "date": e.check_in_date.isoformat(),
                    "group": "Enquiry",
                    "title": "New Enquiry",
                    "property": e.property.title,
                    "location": e.property.city,
                    "guest_name": guest_name,
                    "guest_email": e.email or "",
                    "guest_phone": format_guest_phone(e.country_code, e.phone),
                    "stay": stay_str,
                    "avatar": "",
                    "property_image": "",
                    "status_code": "upcoming",
                    "status_label": "Pending",
                    "guests": f"{e.adults} adult{'' if e.adults == 1 else 's'}{', ' + str(e.children) + ' child' + ('' if e.children == 1 else 'ren') if e.children else ''}",
                    "notes": e.notes_for_host or "",
                })

            # Blocked date events
            blocked_dates = PropertyBlockDate.objects.filter(
                property__created_by__in=get_visible_user_ids(request.user),
                start_date__year=year,
                start_date__month=month,
                is_active=True,
            ).select_related('property')
            for bd in blocked_dates:
                stay_str = ""
                if bd.start_date and bd.end_date:
                    if bd.start_date == bd.end_date:
                        stay_str = f"{bd.start_date:%b %d, %Y}"
                    else:
                        stay_str = f"{bd.start_date:%b %d, %Y}—{bd.end_date:%b %d, %Y}"
                events.append({
                    "id": f"blocked-{bd.id}",
                    "date": bd.start_date.isoformat(),
                    "group": "Blocked",
                    "title": bd.reason or "Blocked",
                    "property": bd.property.title,
                    "location": bd.property.city,
                    "guest_name": "",
                    "guest_email": "",
                    "guest_phone": "",
                    "stay": stay_str,
                    "avatar": "",
                    "property_image": "",
                    "status_code": "blocked",
                    "status_label": "Unavailable",
                })

            events.sort(key=lambda e: e["date"])
            return JsonResponse({"events": events})

        return super().get(request, *args, **kwargs)

@login_required
def channel_integration(request, property_id):
    property_obj = get_object_or_404(Property, pk=property_id, created_by__in=get_visible_user_ids(request.user))

    if request.method == 'POST':
        # --- Handle updates for existing channels ---
        existing_channels = PropertyChannel.objects.filter(property=property_obj)
        for channel in existing_channels:
            calendar_link = request.POST.get(f'calendar_link_{channel.id}')
            is_connected = request.POST.get(f'is_connected_{channel.id}') == 'on'

            # NEW: allow channel type to change when editing
            channel_type_id = request.POST.get(f'channel_type_{channel.id}')
            if channel_type_id:
                try:
                    channel_type_id = int(channel_type_id)
                except (TypeError, ValueError):
                    channel_type_id = None

            # Only update fields that actually changed
            changed = False
            if calendar_link is not None and channel.calendar_link != calendar_link:
                channel.calendar_link = calendar_link
                changed = True

            if channel.is_connected != is_connected:
                channel.is_connected = is_connected
                changed = True

            if channel_type_id and channel.channel_type_id != channel_type_id:
                channel.channel_type_id = channel_type_id
                changed = True

            if changed:
                channel.save()

        # --- Handle new channel additions ---
        new_channel_keys = [k for k in request.POST if k.startswith('new_channel_type_')]
        for key in new_channel_keys:
            timestamp = key.split('_')[-1]
            channel_type_id = request.POST.get(key)
            calendar_link = request.POST.get(f'new_calendar_link_{timestamp}')
            is_connected = request.POST.get(f'new_is_connected_{timestamp}') == 'on'

            if channel_type_id and calendar_link:
                PropertyChannel.objects.update_or_create(
                    property=property_obj,
                    channel_type_id=channel_type_id,
                    defaults={
                        'calendar_link': calendar_link,
                        'is_connected': is_connected
                    }
                )
            elif channel_type_id or calendar_link:
                messages.error(request, "Both channel name and calendar link are required for new integrations.")
                return redirect('channel_integration', property_id=property_id)

        messages.success(request, "Channel integrations saved successfully.")
        return redirect('channel_integration', property_id=property_id)

    channels = PropertyChannel.objects.filter(property=property_obj).select_related('channel_type')
    booking_channels = BookingChannel.objects.all()
    existing_channel_ids = channels.values_list('channel_type_id', flat=True)
    available_booking_channels = booking_channels.exclude(id__in=existing_channel_ids)

    context = {
        'property': property_obj,
        'channels': channels,
        'booking_channels': available_booking_channels,
        'all_booking_channels': booking_channels,
    }
    return render(request, 'frontend/channels/add.html', context)




@login_required
def toggle_channel_status(request, property_id, channel_id):
    """
    API endpoint to toggle the is_connected status of a PropertyChannel.
    """
    if request.method == 'POST':
        try:
            # Ensure the channel belongs to the user and property
            channel = get_object_or_404(PropertyChannel, pk=channel_id, property_id=property_id, property__created_by__in=get_visible_user_ids(request.user))
            
            # Flip the boolean status
            channel.is_connected = not channel.is_connected
            channel.save(update_fields=['is_connected'])

            status_text = "Connected" if channel.is_connected else "Disconnected"
            message = f"Channel '{channel.channel_type.name}' status updated to {status_text}."
            
            return JsonResponse({'success': True, 'is_connected': channel.is_connected, 'message': message})
        except PropertyChannel.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Channel not found.'}, status=404)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@require_POST
def contact_host(request):
    short_code = request.POST.get('host_short_code')
    MyUser = get_user_model()
    host = get_object_or_404(MyUser, short_code=short_code)

    first_name = request.POST.get('first_name')
    last_name = request.POST.get('last_name')
    sender_email = request.POST.get('email')
    phone = request.POST.get('phone')
    message_content = request.POST.get('message')

    subject = f"New Inquiry from {first_name} {last_name} via FavHost"
    email_body = f"""
    Hello {host.get_full_name() or host.username},

    You have received a new message from your public listing page.

    Sender Details:
    - Name: {first_name} {last_name}
    - Email: {sender_email}
    - Phone: {phone}

    Message:
    {message_content}
    """

    try:
        email = EmailMessage(
            subject=subject,
            body=email_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[host.email],
            reply_to=[sender_email],
        )
        email.send(fail_silently=False)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


class ManualSyncAPIView(LoginRequiredMixin, View):
    """
    API endpoint to manually trigger iCal synchronization for all 
    connected channels of the current user.
    """
    def post(self, request, *args, **kwargs):
        from booking.utils import sync_property_channel
        
        # Filter for connected channels on properties owned by the current user
        channels = PropertyChannel.objects.filter(
            property__created_by__in=get_visible_user_ids(request.user), 
            is_connected=True
        )
        count = channels.count()
        for channel in channels:
            sync_property_channel.delay(channel.id)
            
        return JsonResponse({'success': True, 'message': f'Sync started for {count} channels.'})

@login_required
def delete_channel(request, property_id, channel_id):
    """
    API endpoint to delete a PropertyChannel.
    """
    if request.method == 'POST':
        try:
            # Ensure the channel belongs to the user and property
            channel = get_object_or_404(PropertyChannel, pk=channel_id, property_id=property_id, property__created_by__in=get_visible_user_ids(request.user))
            channel_name = channel.channel_type.name
            channel.delete()
            
            return JsonResponse({'success': True, 'message': f"Channel '{channel_name}' deleted successfully."})
        except PropertyChannel.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Channel not found.'}, status=404)
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

class WebsiteIndexView(TemplateView):
    template_name = 'frontend/website/index.html'

class WebsiteBlogView(TemplateView):
    template_name = 'frontend/website/blog.html'

class WebsiteBlogDetailsView(TemplateView):
    def get(self, request, slug, *args, **kwargs):
        # You can update the slug strings below to match your actual blog post slugs.
        if slug == 'the-risks-of-manual-management':
            self.template_name = 'frontend/website/blog-details1.html'
        elif slug == 'the-secret-to-scaling-your-rental-business':   
            self.template_name = 'frontend/website/blog-details2.html'
        elif slug == 'how-to-run-homestay-business-tips-and-tricks':
            self.template_name = 'frontend/website/blog-details3.html'
        elif slug == 'how-to-find-the-right-properties-for-the-short-term-rental-business':
            self.template_name = 'frontend/website/blog-details4.html'
        elif slug == 'how-to-increase-your-visibility-and-ranking-on-booking-com':
            self.template_name = 'frontend/website/blog-details5.html'
        elif slug == 'how-to-increase-your-visibility-and-ranking-on-airbnb':
            self.template_name = 'frontend/website/blog-details6.html'
        else:
            self.template_name = 'frontend/website/blog-details1.html'
            
        return self.render_to_response(self.get_context_data(slug=slug, **kwargs))


class WebsiteContactView(TemplateView):
    template_name = 'frontend/website/contact.html'

    def post(self, request, *args, **kwargs):
        """Handle the "Get in Touch" contact form: email the submission to
        support@favhost.com and show a success/error flash message (PRG)."""
        name = (request.POST.get('name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        subject = (request.POST.get('subject') or '').strip()
        message = (request.POST.get('message') or '').strip()

        if not (name and email and message):
            messages.error(request, 'Please fill in your name, email and message.')
            return redirect('contact')

        to_addr = getattr(settings, 'SUPPORT_EMAIL', 'support@favhost.com')
        mail_subject = f'[Contact form] {subject}' if subject else f'[Contact form] Message from {name}'
        body = (
            f'Name: {name}\n'
            f'Email: {email}\n'
            f'Subject: {subject or "(none)"}\n\n'
            f'{message}\n'
        )

        # Dedicated Zoho SMTP connection for the contact form. Everything else
        # on the platform keeps using the default Namecheap/noreply account.
        support_conn = get_connection(
            backend='django.core.mail.backends.smtp.EmailBackend',
            host=getattr(settings, 'SUPPORT_EMAIL_HOST', None),
            port=getattr(settings, 'SUPPORT_EMAIL_PORT', None),
            username=getattr(settings, 'SUPPORT_EMAIL_HOST_USER', None),
            password=getattr(settings, 'SUPPORT_EMAIL_HOST_PASSWORD', None),
            use_tls=getattr(settings, 'SUPPORT_EMAIL_USE_TLS', True),
        )

        try:
            EmailMessage(
                subject=mail_subject,
                body=body,
                from_email=getattr(settings, 'SUPPORT_EMAIL_HOST_USER', None),
                to=[to_addr],
                reply_to=[email],
                connection=support_conn,
            ).send(fail_silently=False)
            messages.success(request, 'Thanks for reaching out! Your message has been sent to our team.')
        except Exception:
            messages.error(request, 'Sorry, we could not send your message right now. Please email support@favhost.com directly.')

        return redirect('contact')

class WebsiteTermsView(TemplateView):
    template_name = 'frontend/website/Terms&Conditions.html'

class WebsitePrivacyView(TemplateView):
    template_name = 'frontend/website/privacy-policy.html'


# --------------------------------------------------------------------------- #
# Accounting & expense tracking
# --------------------------------------------------------------------------- #
class AccountingView(LoginRequiredMixin, TemplateView):
    """Accounting & expense tracking — income, expenses and profit for a year.

    Income is intentionally computed the SAME way as the rest of the platform
    (Dashboard): the sum of *paid* Payments, by the month of their expected
    payment date, excluding refundable-deposit / refund types. This keeps the
    Accounting figures consistent with the Dashboard and Revenue cards.

    Expenses come from the Expense model (host-recorded). Everything is stored
    in USD (the canonical base) and rendered in the viewer's display currency
    via the {% money %} tag. Profit = income - expenses, per month.
    """
    template_name = 'frontend/accounting/accounting.html'

    MONTH_LABELS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Payment types that are not real income (mirrors Dashboard / Revenue).
    EXCLUDED_PAYMENT_TYPES = ['Refundable deposit', 'Refundable to guest']

    def dispatch(self, request, *args, **kwargs):
        # Co-hosts do not have access to the accountant (accounting) section.
        if request.user.is_authenticated and CoHost.objects.filter(co_host=request.user).exists():
            messages.error(request, 'Accounting is not available for co-hosts.')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        export = request.GET.get('export')
        if export == 'xlsx':
            return self._export_xlsx(request)
        if export == 'csv':
            return self._export_csv(request)
        if request.GET.get('report'):
            return self._render_report(request)
        return super().get(request, *args, **kwargs)

    def _resolve_year(self, user_ids):
        current_year = timezone.now().year
        # Years drawn from both booking activity and recorded expenses.
        booking_years = set(
            Payment.objects
            .filter(booking__property__created_by__in=user_ids, is_paid=True)
            .exclude(type__in=self.EXCLUDED_PAYMENT_TYPES)
            .values_list('expected_payment_date__year', flat=True)
        )
        expense_years = set(
            Expense.objects.filter(created_by__in=user_ids)
            .values_list('date__year', flat=True)
        )
        years = sorted(y for y in (booking_years | expense_years) if y)
        if current_year not in years:
            years = sorted(set(years) | {current_year})

        try:
            selected_year = int(self.request.GET.get('year') or current_year)
        except (TypeError, ValueError):
            selected_year = current_year
        if selected_year not in years:
            selected_year = current_year
        return current_year, selected_year, years

    def _build_report(self, user_ids, selected_year):
        """Return the computed income/expense/profit structure for a year."""
        properties = list(
            Property.objects.filter(created_by__in=user_ids).order_by('title')
        )
        prop_index = {p.id: i for i, p in enumerate(properties)}

        # ---- Income: paid payments by property, by expected-payment month ----
        income_by_prop = [[0.0] * 12 for _ in properties]
        payments = (
            Payment.objects
            .filter(
                booking__property__created_by__in=user_ids,
                is_paid=True,
                expected_payment_date__year=selected_year,
            )
            .exclude(type__in=self.EXCLUDED_PAYMENT_TYPES)
            .values_list('booking__property_id', 'expected_payment_date__month', 'amount')
        )
        for prop_id, month, amount in payments:
            idx = prop_index.get(prop_id)
            if idx is None or not month:
                continue
            income_by_prop[idx][month - 1] += float(amount or 0)

        income_total = [sum(col) for col in zip(*income_by_prop)] if properties else [0.0] * 12
        num_props = len(properties) or 1
        # RevPAR (Revenue Per Available Room, hotel standard): revenue divided by
        # available room-nights, i.e. each property counts as one available unit
        # per night. RevPAR = revenue / (available units × nights in the period).
        days_in_month = [calendar.monthrange(selected_year, m)[1] for m in range(1, 13)]
        revpar = [
            round(income_total[m] / (num_props * days_in_month[m]), 2)
            if days_in_month[m] else 0.0
            for m in range(12)
        ]
        total_nights = num_props * sum(days_in_month)

        income_rows = [{
            'label': p.title,
            'values': income_by_prop[i],
            'overall': sum(income_by_prop[i]),
        } for i, p in enumerate(properties)]

        # ---- Expenses: by category, by month ----
        categories = [c[0] for c in Expense.CATEGORY_CHOICES]
        expense_by_cat = {c: [0.0] * 12 for c in categories}
        for category, month, amount in (
            Expense.objects
            .filter(created_by__in=user_ids, date__year=selected_year)
            .values_list('category', 'date__month', 'amount')
        ):
            if not month:
                continue
            bucket = expense_by_cat.setdefault(category, [0.0] * 12)
            bucket[month - 1] += float(amount or 0)

        expense_rows = [{
            'label': c,
            'values': expense_by_cat[c],
            'overall': sum(expense_by_cat[c]),
        } for c in categories]
        expense_total = [sum(expense_by_cat[c][m] for c in categories) for m in range(12)]

        # ---- Profit = income - expenses ----
        net_profit = [income_total[m] - expense_total[m] for m in range(12)]

        # ---- Profit-by-month bar chart (heights relative to peak magnitude) ----
        # The current calendar month is highlighted orange (only when viewing
        # the current year); loss months render red.
        peak = max((abs(v) for v in net_profit), default=0)
        now = timezone.now()
        current_month_idx = now.month - 1 if now.year == selected_year else None
        bar_cols = [{
            'label': self.MONTH_LABELS[m],
            'value': net_profit[m],
            'height': round(abs(net_profit[m]) / peak * 100, 1) if peak > 0 else 0,
            'negative': net_profit[m] < 0,
            'highlight': m == current_month_idx,
        } for m in range(12)]

        return {
            'properties': properties,
            'income_rows': income_rows,
            'income_total': income_total,
            'income_overall': sum(income_total),
            'revpar': revpar,
            'revpar_overall': round(sum(income_total) / total_nights, 2) if total_nights else 0.0,
            'expense_rows': expense_rows,
            'expense_total': expense_total,
            'expense_overall': sum(expense_total),
            'net_profit': net_profit,
            'net_profit_overall': sum(net_profit),
            'bar_cols': bar_cols,
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_ids = get_visible_user_ids(self.request.user)
        current_year, selected_year, years = self._resolve_year(user_ids)

        report = self._build_report(user_ids, selected_year)

        # ---- Sidebar: recent expenses for the selected year ----
        recent_expenses = (
            Expense.objects
            .filter(created_by__in=user_ids, date__year=selected_year)
            .select_related('property')
        )

        _yidx = years.index(selected_year) if selected_year in years else -1
        context.update(report)
        context.update({
            'month_labels': self.MONTH_LABELS,
            'current_year': current_year,
            'selected_year': selected_year,
            'total_years': years,
            'year_prev': years[_yidx - 1] if _yidx > 0 else None,
            'year_next': years[_yidx + 1] if 0 <= _yidx < len(years) - 1 else None,
            'recent_expenses': recent_expenses,
            'expense_categories': [c[0] for c in Expense.CATEGORY_CHOICES],
            'listings': [{'id': str(p.id), 'title': p.title} for p in report['properties']],
            'has_listings': bool(report['properties']),
        })
        return context

    # Image extensions that can be embedded inline in the printable report;
    # anything else (PDF, docx, …) is shown as a "see attached file" note.
    _IMAGE_EXTS = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg')

    def _render_report(self, request):
        """Render the full, printable detailed expense report for a year.

        This is the "Detailed expense report" option behind the Print button:
        a branded document with a KPI summary, an expense breakdown by category,
        a monthly income/expense/profit summary, a fully itemized expense list
        (every field captured on the Add Expense form) and a receipts appendix
        that embeds each expense's uploaded attachment image.
        """
        from accounts import currency as cur

        user_ids = get_visible_user_ids(request.user)
        current_year, selected_year, years = self._resolve_year(user_ids)
        report = self._build_report(user_ids, selected_year)

        # ---- Itemized expenses (oldest → newest reads naturally in a report) ----
        expenses = list(
            Expense.objects
            .filter(created_by__in=user_ids, date__year=selected_year)
            .select_related('property')
            .order_by('date', 'created_at')
        )

        items = []
        receipts = []
        for e in expenses:
            has_image = False
            att = e.attachment
            if att and att.name:
                has_image = att.name.lower().endswith(self._IMAGE_EXTS)
            receipt_no = None
            if att and att.name:
                receipt_no = len(receipts) + 1
                receipts.append({
                    'no': receipt_no,
                    'expense': e,
                    'is_image': has_image,
                    'filename': att.name.rsplit('/', 1)[-1],
                    'url': att.url,
                })
            items.append({'expense': e, 'receipt_no': receipt_no})

        # ---- Expense breakdown by category (non-zero, largest first, with %) ----
        expense_overall = report['expense_overall'] or 0
        category_breakdown = sorted(
            ({
                'label': row['label'],
                'amount': row['overall'],
                'pct': round(row['overall'] / expense_overall * 100, 1) if expense_overall else 0,
            } for row in report['expense_rows'] if row['overall']),
            key=lambda r: r['amount'], reverse=True,
        )

        # ---- Monthly income / expense / profit summary rows ----
        monthly_summary = [{
            'month': self.MONTH_LABELS[m],
            'income': report['income_total'][m],
            'expense': report['expense_total'][m],
            'profit': report['net_profit'][m],
        } for m in range(12)]

        code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()

        context = {
            'selected_year': selected_year,
            'month_labels': self.MONTH_LABELS,
            'account_name': request.user.get_full_name() or request.user.email,
            'generated_at': timezone.localtime(),
            'display_currency': code,
            'currency_symbol': cur.symbol_for(code),
            'income_overall': report['income_overall'],
            'expense_overall': report['expense_overall'],
            'net_profit_overall': report['net_profit_overall'],
            'expense_count': len(expenses),
            'property_count': len(report['properties']),
            'category_breakdown': category_breakdown,
            'monthly_summary': monthly_summary,
            'items': items,
            'receipts': receipts,
        }
        return render(request, 'frontend/accounting/expense_report.html', context)

    def _export_csv(self, request):
        """Plain CSV of the income / expenses / profit ledger (Excel-ready).

        Money is converted to the viewer's display currency and written as bare
        numbers so spreadsheets treat them as numeric; a UTF-8 BOM is prepended
        so currency symbols render correctly in Excel. Mirrors the Revenue CSV.
        """
        import csv
        import io
        from accounts import currency as cur

        user_ids = get_visible_user_ids(request.user)
        _, selected_year, _ = self._resolve_year(user_ids)
        report = self._build_report(user_ids, selected_year)
        code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()
        symbol = cur.symbol_for(code)
        months = list(self.MONTH_LABELS)
        try:
            generated = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
        except Exception:
            generated = timezone.now().strftime('%Y-%m-%d %H:%M')

        buf = io.StringIO()
        buf.write('﻿')  # UTF-8 BOM so Excel renders ₹/€/₨ etc. correctly
        writer = csv.writer(buf)

        writer.writerow(['FavHost - Accounting & Expense Tracking'])
        writer.writerow(['Year', selected_year])
        writer.writerow(['Currency', f'{code} ({symbol})'])
        writer.writerow(['Generated', generated])
        writer.writerow([])

        def money(value):
            return cur.money_raw(value, code)

        def section(title, first_header, rows):
            writer.writerow([title])
            writer.writerow([first_header] + months + ['Overall'])
            for label, values, overall in rows:
                writer.writerow([label] + [money(v) for v in values] + [money(overall)])
            writer.writerow([])

        income_rows = [(row['label'], row['values'], row['overall']) for row in report['income_rows']]
        income_rows.append(('Total', report['income_total'], report['income_overall']))
        income_rows.append(('RevPAR', report['revpar'], report['revpar_overall']))
        section('Income', 'Property', income_rows)

        expense_rows = [(row['label'], row['values'], row['overall']) for row in report['expense_rows']]
        expense_rows.append(('Total', report['expense_total'], report['expense_overall']))
        section('Expenses', 'Category', expense_rows)

        section('Profit', 'Metric',
                [('Net Profit', report['net_profit'], report['net_profit_overall'])])

        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="accounting_{selected_year}.csv"'
        return response

    def _export_xlsx(self, request):
        """Styled .xlsx of the income / expenses / profit ledger for the year."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from accounts import currency as cur

        user_ids = get_visible_user_ids(request.user)
        _, selected_year, _ = self._resolve_year(user_ids)
        report = self._build_report(user_ids, selected_year)
        code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()
        symbol = cur.symbol_for(code)
        months = self.MONTH_LABELS
        ncols = 1 + 12 + 1

        # Palette mirrors the on-screen page: brand orange accents,
        # charcoal section bands, warm-orange overall/total cells.
        title_font = Font(bold=True, size=14, color='EB5310')
        meta_label_font = Font(bold=True, color='6B7280')
        section_font = Font(bold=True, color='FFFFFF')
        section_fill = PatternFill('solid', fgColor='313131')
        header_font = Font(bold=True, color='1A1A1A')
        header_fill = PatternFill('solid', fgColor='FAF7F5')
        overall_fill = PatternFill('solid', fgColor='FFF3EC')
        overall_font = Font(bold=True, color='C2410C')
        label_font = Font(bold=True, color='1A1A1A')
        label_fill = PatternFill('solid', fgColor='FAF7F5')
        total_font = Font(bold=True, color='C2410C')
        total_fill = PatternFill('solid', fgColor='FFF3EC')
        MONEY_FMT = '#,##0.00'
        thin = Side(style='thin', color='E5E7EB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Accounting'

        r = 1
        ws.cell(r, 1, 'FavHost - Accounting & Expense Tracking').font = title_font
        r += 1
        try:
            generated = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
        except Exception:
            generated = timezone.now().strftime('%Y-%m-%d %H:%M')
        for label, val in [('Year', selected_year), ('Currency', f'{code} ({symbol})'), ('Generated', generated)]:
            ws.cell(r, 1, label).font = meta_label_font
            ws.cell(r, 2, val)
            r += 1
        r += 1

        def money_cell(row_i, col_i, usd):
            c = ws.cell(row_i, col_i)
            c.value = float(cur.money_raw(usd, code))
            c.number_format = MONEY_FMT
            c.border = border
            c.alignment = Alignment(horizontal='right')
            return c

        def section_header(title):
            nonlocal r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            tc = ws.cell(r, 1, title)
            tc.font = section_font
            tc.fill = section_fill
            tc.alignment = Alignment(horizontal='center')
            r += 1
            for i, h in enumerate(['' ] + months + ['Overall'], start=1):
                hc = ws.cell(r, i, h)
                hc.font = overall_font if h == 'Overall' else header_font
                hc.fill = overall_fill if h == 'Overall' else header_fill
                hc.border = border
                hc.alignment = Alignment(horizontal='left' if i == 1 else 'center')
            r += 1

        def data_row(label, values, overall, bold=False):
            nonlocal r
            lc = ws.cell(r, 1, label)
            lc.font = total_font if bold else label_font
            lc.fill = total_fill if bold else label_fill
            lc.border = border
            for i, v in enumerate(list(values), start=2):
                cell = money_cell(r, i, v)
                if bold:
                    cell.fill = total_fill
                    cell.font = total_font
            oc = money_cell(r, ncols, overall)
            oc.fill = total_fill if bold else overall_fill
            oc.font = total_font if bold else overall_font
            r += 1

        section_header('Income')
        for row in report['income_rows']:
            data_row(row['label'], row['values'], row['overall'])
        data_row('Total', report['income_total'], report['income_overall'], bold=True)
        data_row('RevPAR', report['revpar'], report['revpar_overall'])
        r += 1

        section_header('Expenses')
        for row in report['expense_rows']:
            data_row(row['label'], row['values'], row['overall'])
        data_row('Total', report['expense_total'], report['expense_overall'], bold=True)
        r += 1

        section_header('Profit')
        data_row('Net Profit', report['net_profit'], report['net_profit_overall'], bold=True)

        ws.column_dimensions['A'].width = 24
        for i in range(2, ncols):
            ws.column_dimensions[get_column_letter(i)].width = 11
        ws.column_dimensions[get_column_letter(ncols)].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="accounting_{selected_year}.xlsx"'
        return response


def _parse_expense_amount_to_usd(request, raw_amount):
    """Convert a host-entered amount (display currency) to the USD base."""
    from accounts import currency as cur
    code = (getattr(request.user, 'currency', None) or cur.BASE_CURRENCY).upper()
    return cur.to_usd(raw_amount or 0, code)


def _redirect_to_accounting(request):
    """Return to the accounting page, preserving the selected year."""
    year = request.POST.get('year') or request.GET.get('year')
    url = '/accounting/'
    if year:
        url = f'{url}?year={year}'
    return redirect(url)


def _cohost_blocked(request):
    """Redirect co-hosts away from accountant actions; returns None otherwise."""
    if request.user.is_authenticated and CoHost.objects.filter(co_host=request.user).exists():
        messages.error(request, 'Accounting is not available for co-hosts.')
        return redirect('dashboard')
    return None


def _redirect_after_add_expense(request):
    """Return the user to the right place after adding an expense.

    Co-hosts add expenses from the shared modal in the header (they cannot open
    /accounting/), so they go back to the page they came from. Hosts return to
    the Accounting page (preserving the year), exactly as before.
    """
    if CoHost.objects.filter(co_host=request.user).exists():
        referer = request.META.get('HTTP_REFERER')
        return redirect(referer) if referer else redirect('dashboard')
    return _redirect_to_accounting(request)


@login_required
@require_POST
def add_expense(request):
    # Note: co-hosts ARE allowed to add expenses (but not to view the
    # Accounting/Revenue pages, which are guarded separately).
    user_ids = get_visible_user_ids(request.user)
    category = request.POST.get('category') or 'Other expenses'
    date_str = request.POST.get('date')
    note = request.POST.get('note') or ''
    property_id = request.POST.get('property')

    if not date_str:
        messages.error(request, 'Please provide a date for the expense.')
        return _redirect_after_add_expense(request)

    try:
        amount_usd = _parse_expense_amount_to_usd(request, request.POST.get('amount'))
    except Exception:
        messages.error(request, 'Please enter a valid amount.')
        return _redirect_after_add_expense(request)

    prop = None
    if property_id:
        prop = Property.objects.filter(id=property_id, created_by__in=user_ids).first()

    Expense.objects.create(
        created_by=request.user,
        property=prop,
        category=category,
        amount=amount_usd,
        date=date_str,
        note=note,
        attachment=request.FILES.get('attachment'),
    )
    messages.success(request, 'Expense added successfully.')
    return _redirect_after_add_expense(request)


@login_required
@require_POST
def edit_expense(request, pk):
    blocked = _cohost_blocked(request)
    if blocked:
        return blocked
    user_ids = get_visible_user_ids(request.user)
    expense = get_object_or_404(Expense, pk=pk, created_by__in=user_ids)

    expense.category = request.POST.get('category') or expense.category
    date_str = request.POST.get('date')
    if date_str:
        expense.date = date_str
    expense.note = request.POST.get('note') or ''

    amount_raw = request.POST.get('amount')
    if amount_raw not in (None, ''):
        try:
            expense.amount = _parse_expense_amount_to_usd(request, amount_raw)
        except Exception:
            messages.error(request, 'Please enter a valid amount.')
            return _redirect_to_accounting(request)

    property_id = request.POST.get('property')
    if property_id:
        expense.property = Property.objects.filter(id=property_id, created_by__in=user_ids).first()
    else:
        expense.property = None

    if request.FILES.get('attachment'):
        expense.attachment = request.FILES['attachment']

    expense.save()
    messages.success(request, 'Expense updated successfully.')
    return _redirect_to_accounting(request)


@login_required
@require_POST
def delete_expense(request, pk):
    blocked = _cohost_blocked(request)
    if blocked:
        return blocked
    user_ids = get_visible_user_ids(request.user)
    expense = get_object_or_404(Expense, pk=pk, created_by__in=user_ids)
    expense.delete()
    messages.success(request, 'Expense deleted successfully.')
    return _redirect_to_accounting(request)
